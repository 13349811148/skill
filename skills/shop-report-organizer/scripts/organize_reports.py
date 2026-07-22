#!/usr/bin/env python
"""Organize shop backend reports by reading spreadsheet content."""

from __future__ import annotations

import argparse
import codecs
import csv
import datetime as dt
import hashlib
import json
import math
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import zipfile
import zlib
from collections import Counter
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Iterable


SKILL_DIR = Path(__file__).resolve().parents[1]
REFERENCES_DIR = SKILL_DIR / "references"
SIGNATURES_PATH = REFERENCES_DIR / "report_signatures.csv"
PRODUCT_MAP_PATH = REFERENCES_DIR / "shop_product_map.csv"
EXPECTED_REPORTS_PATH = REFERENCES_DIR / "expected_reports.csv"
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
SUPPORTED_ARCHIVE_EXTENSIONS = {".zip"}
EXTRACTED_ZIP_FOLDER_NAME = "_已解压ZIP"
DUPLICATE_INPUT_FOLDER_NAME = "_重复文件"
INVALID_FILENAME_CHARS = r'<>:"/\|?*'
PRODUCT_ID_CHECK_REPORT_TYPES = {"订单数据", "售后数据", "推广数据"}
PRODUCT_DATA_REQUIRED_SOURCE_REPORT_TYPES = {"订单数据", "推广数据"}
PRODUCT_HISTORY_SOURCE_REPORT_TYPES = {"售后数据"}
ORDER_REPORT_TYPE = "订单数据"
ORDER_DATA_PREFIX = "订单数据"
AFTER_SALE_REPORT_TYPE = "售后数据"
AFTER_SALE_UPDATE_PREFIX = "售后数据更新至"
PROMOTION_ADJUSTMENT_REPORT_TYPE = "推广调整日志"
PROMOTION_ADJUSTMENT_UPDATE_PREFIX = "推广调整日志更新至"
ORDER_SUMMARY_PREFIX = "订单汇总表"
ORDER_ID_HEADERS = ["订单号", "子订单编号", "主订单编号"]
ORDER_PAYMENT_TIME_HEADERS = [
    "订单支付时间",
    "支付时间",
    "订单付款时间",
]
ORDER_TRANSACTION_TIME_HEADERS = ["订单成交时间"]
ORDER_CREATION_TIME_HEADERS = ["订单创建时间"]
ORDER_TIME_HEADERS = [
    *ORDER_PAYMENT_TIME_HEADERS,
    *ORDER_TRANSACTION_TIME_HEADERS,
    *ORDER_CREATION_TIME_HEADERS,
]
PREVIEW_MAX_ROWS = 100_000
EXCEL_MAX_DATA_ROWS = 1_048_575
STREAM_COMMIT_INTERVAL = 2_000
TMALL_PLATFORM = "天猫"
TMALL_SHOP_SUFFIX = "（天猫）"
ORDER_SUMMARY_METADATA_HEADERS = ["汇总_店铺名称", "汇总_最新下载时间", "汇总_最新来源文件", "汇总_最新归档文件"]
PROTECTED_ORDER_INFO_EXACT_HEADERS = {"消费者资料", "省", "市", "区", "用户购买手机号", "门店名称", "门店自定义编码"}
PROTECTED_ORDER_INFO_KEYWORDS = ["收件", "收货地址", "收货人", "详细地址", "地址", "手机号", "电话"]
PROTECTED_ORDER_INFO_EXCLUDE_KEYWORDS = ["仓库", "配送员", "时间", "状态", "是否"]
MISSING_REMINDER_FIELDNAMES = [
    "action",
    "status",
    "report_type",
    "detected_count",
    "shop_name",
    "source_name",
    "source_path",
    "new_product_ids",
    "message",
    "input_path",
]


@dataclass
class Signature:
    report_type: str
    target_folder: str
    platform: str
    required_headers: list[str]
    product_id_headers: list[str]
    period_headers: list[str]
    period_mode: str
    notes: str


@dataclass
class Analysis:
    source: Path
    status: str
    reason: str
    report_type: str
    target_folder: str
    shop_name: str
    shop_source: str
    period: str
    period_source: str
    download_date: str
    target_path: Path
    product_counts: str
    matched_headers: str
    missing_fields: str
    map_updates: list[tuple[str, str]]
    product_ids: list[str]
    new_product_ids: list[str]


@dataclass
class ExpectedReport:
    report_type: str
    required: bool
    reminder_message: str


@dataclass
class OrderFileScan:
    month_keys: set[str]
    read_rows: int
    valid_rows: int
    ignored_rows: int


def read_utf8_sig_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_utf8_sig_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def split_cell(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split("|") if part.strip()]


def load_signatures(path: Path = SIGNATURES_PATH) -> list[Signature]:
    signatures: list[Signature] = []
    for row in read_utf8_sig_csv(path):
        signatures.append(
            Signature(
                report_type=row["report_type"].strip(),
                target_folder=row["target_folder"].strip(),
                platform=normalize_text(row.get("platform", "")),
                required_headers=split_cell(row.get("required_headers")),
                product_id_headers=split_cell(row.get("product_id_headers")),
                period_headers=split_cell(row.get("period_headers")),
                period_mode=row.get("period_mode", "").strip(),
                notes=row.get("notes", "").strip(),
            )
        )
    return signatures


def truthy(value: object) -> bool:
    text = normalize_text(value).lower()
    return text in {"1", "true", "yes", "y", "是", "必备", "required"}


def load_expected_reports(path: Path = EXPECTED_REPORTS_PATH) -> list[ExpectedReport]:
    if not path.exists():
        return []
    expected_reports: list[ExpectedReport] = []
    for row in read_utf8_sig_csv(path):
        report_type = normalize_text(row.get("report_type", ""))
        if not report_type:
            continue
        expected_reports.append(
            ExpectedReport(
                report_type=report_type,
                required=truthy(row.get("required", "")),
                reminder_message=normalize_text(row.get("reminder_message", "")),
            )
        )
    return expected_reports


def load_product_map(path: Path = PRODUCT_MAP_PATH) -> dict[str, str]:
    if not path.exists():
        return {}
    mapping: dict[str, str] = {}
    for row in read_utf8_sig_csv(path):
        product_id = normalize_product_id(row.get("product_id", ""))
        shop_name = (row.get("shop_name") or "").strip()
        if product_id and shop_name:
            mapping[product_id] = shop_name
    return mapping


def save_product_map(mapping: dict[str, str], path: Path = PRODUCT_MAP_PATH) -> None:
    rows = [
        {"product_id": product_id, "shop_name": shop_name, "source": "organize_reports"}
        for product_id, shop_name in sorted(mapping.items(), key=lambda item: (item[1], item[0]))
    ]
    write_utf8_sig_csv(path, rows, ["product_id", "shop_name", "source"])


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def normalize_header(value: object) -> str:
    return normalize_text(value).lower().replace(" ", "")


def normalize_product_id(value: object) -> str:
    text = normalize_text(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) >= 6 else ""


def read_csv_rows(path: Path, max_rows: int) -> list[list[str]]:
    encoding = detect_csv_encoding(path)
    rows: list[list[str]] = []
    with path.open("r", encoding=encoding, errors="strict", newline="") as handle:
        first_line = handle.readline()
        delimiter = "\t" if "\t" in first_line and "," not in first_line else ","
        handle.seek(0)
        for index, row in enumerate(csv.reader(handle, delimiter=delimiter)):
            if index >= max_rows:
                break
            values = [normalize_text(cell) for cell in row]
            if any(values):
                rows.append(values)
    return rows


def read_xlsx_rows(path: Path, max_rows: int, max_cols: int) -> list[list[str]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment specific
        raise RuntimeError("openpyxl is required for .xlsx files") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[list[str]] = []
    # Some platform exports keep their usable headers in a hidden worksheet.
    # Inspect all worksheets while keeping the per-file row limit intact.
    for worksheet in workbook.worksheets:
        if len(rows) >= max_rows:
            break
        if hasattr(worksheet, "reset_dimensions"):
            try:
                dimension = worksheet.calculate_dimension()
            except ValueError:
                worksheet.reset_dimensions()
            else:
                if dimension == "A1:A1":
                    worksheet.reset_dimensions()
        remaining_rows = max_rows - len(rows)
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=remaining_rows,
            max_col=max_cols,
            values_only=True,
        ):
            values = [normalize_text(cell) for cell in row]
            if any(values):
                rows.append(values)
    workbook.close()
    return rows


def powershell_exe() -> str:
    return shutil.which("powershell.exe") or shutil.which("powershell") or "powershell"


def read_xls_rows(path: Path, max_rows: int, max_cols: int) -> list[list[str]]:
    ps_script = rf"""
$ErrorActionPreference = 'Stop'
[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false)
$path = @'
{str(path)}
'@
$maxRows = {max_rows}
$maxCols = {max_cols}
$excel = $null
try {{
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $wb = $excel.Workbooks.Open($path, 0, $true)
  try {{
    $ws = $wb.Worksheets.Item(1)
    $used = $ws.UsedRange
    $rows = [Math]::Min($maxRows, $used.Rows.Count)
    $cols = [Math]::Min($maxCols, $used.Columns.Count)
    $result = @()
    for ($r = 1; $r -le $rows; $r++) {{
      $vals = @()
      for ($c = 1; $c -le $cols; $c++) {{
        $vals += [string]$ws.Cells.Item($r, $c).Text
      }}
      if ((($vals -join '')).Trim().Length -gt 0) {{
        $result += [PSCustomObject]@{{ cells = $vals }}
      }}
    }}
    $result | ConvertTo-Json -Compress -Depth 5
  }} finally {{
    $wb.Close($false)
  }}
}} finally {{
  if ($excel -ne $null) {{
    $excel.Quit() | Out-Null
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
  }}
}}
"""
    completed = subprocess.run(
        [powershell_exe(), "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        encoding="utf-8",
        errors="replace",
    )
    output = completed.stdout.strip()
    if not output:
        return []
    parsed = json.loads(output)
    if isinstance(parsed, dict):
        parsed = [parsed]
    rows = [item.get("cells", []) for item in parsed if isinstance(item, dict)]
    return [[normalize_text(cell) for cell in row] for row in rows if any(normalize_text(cell) for cell in row)]


def read_rows(path: Path, max_rows: int, max_cols: int) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path, max_rows)
    if suffix == ".xlsx":
        return read_xlsx_rows(path, max_rows, max_cols)
    if suffix == ".xls":
        return read_xls_rows(path, max_rows, max_cols)
    raise RuntimeError(f"unsupported extension: {suffix}")


def detect_csv_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        decoder = codecs.getincrementaldecoder(encoding)(errors="strict")
        try:
            with path.open("rb") as handle:
                while chunk := handle.read(1024 * 1024):
                    decoder.decode(chunk)
                decoder.decode(b"", final=True)
            return encoding
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"CSV encoding cannot be decoded: {path}")


def iter_csv_rows_full(path: Path) -> Iterable[list[str]]:
    encoding = detect_csv_encoding(path)
    with path.open("r", encoding=encoding, errors="strict", newline="") as handle:
        first_line = handle.readline()
        delimiter = "\t" if "\t" in first_line and "," not in first_line else ","
        handle.seek(0)
        for row in csv.reader(handle, delimiter=delimiter):
            values = [normalize_text(cell) for cell in row]
            if any(values):
                yield values


def iter_xlsx_rows_full(path: Path, max_cols: int) -> Iterable[list[str]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment specific
        raise RuntimeError("openpyxl is required for .xlsx files") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            if hasattr(worksheet, "reset_dimensions"):
                try:
                    dimension = worksheet.calculate_dimension()
                except ValueError:
                    worksheet.reset_dimensions()
                else:
                    if dimension == "A1:A1":
                        worksheet.reset_dimensions()
            for row in worksheet.iter_rows(min_row=1, max_col=max_cols, values_only=True):
                values = [normalize_text(cell) for cell in row]
                if any(values):
                    yield values
    finally:
        workbook.close()


def iter_rows_full(path: Path, max_cols: int) -> Iterable[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv_rows_full(path)
        return
    if suffix == ".xlsx":
        yield from iter_xlsx_rows_full(path, max_cols)
        return
    if suffix == ".xls":
        # Legacy .xls worksheets cannot exceed 65,536 rows, so a complete COM
        # read remains below the Excel single-sheet limit.
        yield from read_xls_rows(path, EXCEL_MAX_DATA_ROWS + 1, max_cols)
        return
    raise RuntimeError(f"unsupported extension: {suffix}")


def header_contains(headers: list[str], required: str) -> bool:
    needle = normalize_header(required)
    return any(needle in normalize_header(header) for header in headers)


def find_header_index(headers: list[str], candidates: Iterable[str]) -> int | None:
    for candidate in candidates:
        needle = normalize_header(candidate)
        for index, header in enumerate(headers):
            if needle and needle in normalize_header(header):
                return index
    return None


def find_header_indices(headers: list[str], candidates: Iterable[str]) -> list[tuple[str, int]]:
    matches: list[tuple[str, int]] = []
    seen_indices: set[int] = set()
    for candidate in candidates:
        needle = normalize_header(candidate)
        index = next(
            (
                header_index
                for header_index, header in enumerate(headers)
                if header_index not in seen_indices and normalize_header(header) == needle
            ),
            None,
        )
        if index is None:
            index = next(
                (
                    header_index
                    for header_index, header in enumerate(headers)
                    if header_index not in seen_indices
                    and needle
                    and needle in normalize_header(header)
                ),
                None,
            )
        if index is not None and index not in seen_indices:
            matches.append((candidate, index))
            seen_indices.add(index)
    return matches


def order_time_columns(headers: list[str]) -> list[tuple[str, int]]:
    candidates = [*ORDER_PAYMENT_TIME_HEADERS, *ORDER_TRANSACTION_TIME_HEADERS]
    if (
        find_header_index(headers, ["子订单编号"]) is not None
        and find_header_index(headers, ["主订单编号"]) is not None
    ):
        candidates.extend(ORDER_CREATION_TIME_HEADERS)
    return find_header_indices(headers, candidates)


def match_signature(rows: list[list[str]], signatures: list[Signature]) -> tuple[Signature | None, int, list[str], list[str]]:
    best: tuple[int, Signature | None, int, list[str], list[str]] = (-1, None, -1, [], [])
    for row_index, row in enumerate(rows[:20]):
        headers = row
        for signature in signatures:
            matched = [header for header in signature.required_headers if header_contains(headers, header)]
            if len(matched) == len(signature.required_headers) and len(matched) > best[0]:
                best = (len(matched), signature, row_index, headers, matched)
    return best[1], best[2], best[3], best[4]


def collect_product_ids(rows: list[list[str]], header_row_index: int, headers: list[str], signature: Signature | None) -> list[str]:
    candidates = signature.product_id_headers if signature else ["商品ID", "商品id"]
    column_index = find_header_index(headers, candidates)
    if column_index is None:
        for index, header in enumerate(headers):
            normalized = normalize_header(header)
            if "商品" in normalized and "id" in normalized:
                column_index = index
                break
    product_ids: list[str] = []
    if column_index is not None:
        for row in rows[header_row_index + 1 :]:
            if column_index < len(row):
                product_id = normalize_product_id(row[column_index])
                if product_id:
                    product_ids.append(product_id)
    if not product_ids:
        product_id_pattern = re.compile(r"\u5546\u54c1\s*ID[\uff1a:]\s*(\d{6,})", flags=re.IGNORECASE)
        for row in rows[header_row_index + 1 :]:
            for cell in row:
                for match in product_id_pattern.finditer(normalize_text(cell)):
                    product_id = normalize_product_id(match.group(1))
                    if product_id:
                        product_ids.append(product_id)
    return product_ids


def shop_from_product_archive_path(path: Path, database_root: Path) -> str:
    product_root = database_root / "商品数据"
    try:
        relative = path.relative_to(product_root)
    except ValueError:
        return ""
    if not relative.parts:
        return ""
    shop_name = relative.parts[0]
    return "" if shop_name.startswith("_") else shop_name


def load_archived_product_map(
    database_root: Path,
    signatures: list[Signature],
    existing_map: dict[str, str],
    max_rows: int,
    max_cols: int,
) -> dict[str, str]:
    product_root = database_root / "商品数据"
    if not product_root.exists():
        return {}

    archived_map: dict[str, str] = {}
    for path in sorted(product_root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        shop_name = shop_from_product_archive_path(path, database_root)
        if not shop_name:
            continue
        try:
            rows = read_rows(path, max_rows, max_cols)
            signature, header_row_index, headers, _matched = match_signature(rows, signatures)
        except Exception:
            continue
        if signature is None or signature.report_type != "商品数据":
            continue
        for product_id in collect_product_ids(rows, header_row_index, headers, signature):
            if product_id and product_id not in existing_map and product_id not in archived_map:
                archived_map[product_id] = shop_name
    return archived_map


def known_shops(database_root: Path, product_map: dict[str, str]) -> list[str]:
    shops = set(product_map.values())
    for report_root in ("订单数据", "售后数据", "商品数据", "推广数据", "推广调整日志"):
        root = database_root / report_root
        if root.exists():
            for child in root.iterdir():
                if child.is_dir() and not child.name.startswith("_"):
                    shops.add(child.name)
    return sorted(shops, key=len, reverse=True)


def strip_platform_suffix(shop_name: str) -> str:
    return re.sub(r"[（(]\s*天猫\s*[）)]$", "", normalize_text(shop_name), flags=re.IGNORECASE)


def is_tmall_shop(shop_name: str) -> bool:
    return bool(re.search(r"[（(]\s*天猫\s*[）)]$", normalize_text(shop_name), flags=re.IGNORECASE))


def normalize_shop_platform(shop_name: str, platform: str) -> str:
    shop_name = normalize_text(shop_name)
    if not shop_name:
        return ""
    if normalize_text(platform).lower() in {"天猫", "tmall"}:
        return f"{strip_platform_suffix(shop_name)}{TMALL_SHOP_SUFFIX}"
    return shop_name


def detect_platform(path: Path, rows: list[list[str]], signature: Signature) -> tuple[str, str]:
    if signature.platform:
        return signature.platform, "report_signature"
    searchable = [path.name]
    for row in rows[:300]:
        searchable.extend(normalize_text(cell) for cell in row if normalize_text(cell))
    text = " ".join(searchable).lower()
    if "天猫" in text or "tmall" in text:
        return TMALL_PLATFORM, "filename_or_spreadsheet_text"
    return "", ""


def canonicalize_shop(text: str, shops: list[str]) -> str:
    compact = re.sub(r"\s+", "", text)
    text_is_tmall = "天猫" in compact or "tmall" in compact.lower()
    ordered_shops = sorted(shops, key=lambda shop: (is_tmall_shop(shop) != text_is_tmall, -len(shop)))
    for shop in ordered_shops:
        base_shop = strip_platform_suffix(shop)
        if shop and shop in compact:
            return shop
        if base_shop and base_shop in compact:
            return shop
        short_shop = re.sub(r"(旗舰店|专卖店|专营店)$", "", base_shop)
        if short_shop and short_shop in compact:
            return shop
    match = re.search(r"商家[（(]([^）)]+)[）)]", compact)
    if match:
        candidate = re.sub(r"(客服|平台|商家|极速退款)", "", match.group(1))
        candidate = re.sub(r"(\d+号|[A-Za-z]*\d+|晶晶|小惠)$", "", candidate)
        for shop in ordered_shops:
            base_shop = strip_platform_suffix(shop)
            if candidate and (candidate in base_shop or base_shop in candidate):
                return shop
    return ""


def extract_explicit_shop(rows: list[list[str]], shops: list[str]) -> tuple[str, str]:
    for row in rows[:300]:
        for cell in row:
            text = normalize_text(cell)
            if not text:
                continue
            shop = canonicalize_shop(text, shops)
            if shop:
                return shop, "spreadsheet_text"
    return "", ""


def extract_shop_from_filename(path: Path, shops: list[str]) -> tuple[str, str]:
    match = re.search(r"[（(]([^）)]+)[）)]", path.name)
    if not match:
        return "", ""
    text = match.group(1).strip()
    if re.fullmatch(r"\d+", text):
        return "", ""
    canonical_shop = canonicalize_shop(text, shops)
    if canonical_shop:
        return canonical_shop, "filename_parentheses"
    text = re.sub(r"^(?:天猫|淘宝|拼多多|pdd)\s*[-—_ ]+", "", text, flags=re.IGNORECASE)
    for shop in shops:
        if text == shop or text in shop or shop in text:
            return shop, "filename_parentheses"
    return text, "filename_parentheses"


def decide_shop(
    path: Path,
    report_type: str,
    rows: list[list[str]],
    product_ids: list[str],
    product_map: dict[str, str],
    shops: list[str],
) -> tuple[str, str, str]:
    explicit_shop, explicit_source = extract_explicit_shop(rows, shops)
    if explicit_shop:
        return explicit_shop, explicit_source, ""

    counts = Counter(product_map[product_id] for product_id in product_ids if product_id in product_map)
    counts_text = ";".join(f"{shop}:{count}" for shop, count in counts.most_common())
    if counts:
        top_shop, top_count = counts.most_common(1)[0]
        second_count = counts.most_common(2)[1][1] if len(counts) > 1 else 0
        if top_count > second_count:
            return top_shop, "product_id_map", counts_text

    if report_type == "商品数据":
        filename_shop, source = extract_shop_from_filename(path, shops)
        if filename_shop:
            return filename_shop, source, counts_text

    return "", "", counts_text


DATE_PATTERNS = [
    re.compile(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})"),
    re.compile(r"\b(20\d{2})(\d{2})(\d{2})\b"),
    re.compile(r"\b(\d{2})(\d{2})(\d{2})\b"),
]


def parse_date(text: object) -> dt.date | None:
    value = normalize_text(text)
    for pattern in DATE_PATTERNS:
        match = pattern.search(value)
        if not match:
            continue
        parts = match.groups()
        if len(parts[0]) == 2:
            year = 2000 + int(parts[0])
        else:
            year = int(parts[0])
        try:
            return dt.date(year, int(parts[1]), int(parts[2]))
        except ValueError:
            continue
    return None


def date_range_from_column(rows: list[list[str]], header_row_index: int, headers: list[str], period_headers: list[str]) -> tuple[str, str] | None:
    columns = find_header_indices(headers, period_headers)
    if not columns:
        return None
    dates: list[dt.date] = []
    for row in rows[header_row_index + 1 :]:
        for _header, column_index in columns:
            parsed = parse_date(row[column_index]) if column_index < len(row) else None
            if parsed:
                dates.append(parsed)
                break
    if not dates:
        return None
    start = min(dates).isoformat()
    end = max(dates).isoformat()
    return start, end


def date_range_from_filename(path: Path) -> tuple[str, str] | None:
    text = path.name
    match = re.search(r"(20\d{6})\s*至\s*(20\d{6})", text)
    if match:
        start = parse_date(match.group(1))
        end = parse_date(match.group(2))
        if start and end:
            return start.isoformat(), end.isoformat()
    match = re.search(r"(20\d{2}[-.]\d{1,2}[-.]\d{1,2})\s*至\s*(20\d{2}[-.]\d{1,2}[-.]\d{1,2})", text)
    if match:
        start = parse_date(match.group(1))
        end = parse_date(match.group(2))
        if start and end:
            return start.isoformat(), end.isoformat()
    return None


def file_creation_datetime(path: Path) -> dt.datetime:
    stat_result = path.stat()
    if hasattr(stat_result, "st_birthtime"):
        timestamp = stat_result.st_birthtime
    elif os.name == "nt":
        timestamp = stat_result.st_ctime
    else:
        timestamp = stat_result.st_mtime
    return dt.datetime.fromtimestamp(timestamp)


def download_date(path: Path) -> dt.date:
    return file_creation_datetime(path).date()


def download_datetime(path: Path) -> dt.datetime:
    return file_creation_datetime(path)


def decide_period(
    path: Path,
    signature: Signature,
    rows: list[list[str]],
    header_row_index: int,
    headers: list[str],
    file_download_date: dt.date,
) -> tuple[str, str]:
    if signature.period_mode == "download_date":
        return file_download_date.isoformat(), "download_date"

    if signature.period_mode == "filename_range_then_date_column":
        column_range = date_range_from_column(rows, header_row_index, headers, signature.period_headers)
        if "分天数据" in normalize_text(path.name) and column_range and column_range[0] != column_range[1]:
            return f"{column_range[0]}至{column_range[1]}", "date_column"
        filename_range = date_range_from_filename(path)
        if filename_range:
            return f"{filename_range[0]}至{filename_range[1]}", "filename_range"
        if column_range:
            return f"{column_range[0]}至{column_range[1]}", "date_column"
        return "", ""

    column_range = date_range_from_column(rows, header_row_index, headers, signature.period_headers)
    if column_range:
        return f"{column_range[0]}至{column_range[1]}", "date_column"
    return "", ""


def period_date_range(period: str) -> tuple[str, str] | None:
    match = re.fullmatch(r"(20\d{2}-\d{2}-\d{2})至(20\d{2}-\d{2}-\d{2})", normalize_text(period))
    if not match:
        return None
    return match.group(1), match.group(2)


def is_daily_breakdown_promotion(
    report_type: str,
    rows: list[list[str]],
    header_row_index: int,
    headers: list[str],
    period_headers: list[str],
    product_id_headers: list[str],
) -> bool:
    # Whether a promotion export can be archived is determined by its data,
    # not by a filename convention. Tmall exports often combine daily rows
    # without including "分天数据" in the filename.
    if report_type != "推广数据":
        return False

    date_column_index = find_header_index(headers, period_headers)
    product_id_column_index = find_header_index(headers, product_id_headers)
    if date_column_index is None or product_id_column_index is None:
        return False

    dates: list[dt.date] = []
    data_row_count = 0
    for row in rows[header_row_index + 1 :]:
        if not any(normalize_text(value) for value in row):
            continue
        first_cell = normalize_text(row[0]) if row else ""
        if first_cell in {"总计", "合计"} or first_cell.startswith("注"):
            continue
        data_row_count += 1
        if date_column_index >= len(row):
            return False
        parsed = parse_date(row[date_column_index])
        if not parsed:
            return False
        if product_id_column_index >= len(row) or not normalize_product_id(row[product_id_column_index]):
            return False
        dates.append(parsed)

    return data_row_count > 0 and len(set(dates)) > 1


def is_multi_day_promotion_summary(
    path: Path,
    report_type: str,
    period: str,
    rows: list[list[str]],
    header_row_index: int,
    headers: list[str],
    period_headers: list[str],
    product_id_headers: list[str],
) -> bool:
    if report_type != "推广数据":
        return False
    date_range = period_date_range(period)
    if not date_range or date_range[0] == date_range[1]:
        return False
    return not is_daily_breakdown_promotion(
        report_type,
        rows,
        header_row_index,
        headers,
        period_headers,
        product_id_headers,
    )


def sanitize_filename_part(value: str) -> str:
    value = normalize_text(value)
    for char in INVALID_FILENAME_CHARS:
        value = value.replace(char, "_")
    value = re.sub(r"\s+", " ", value).strip(" .")
    return value or "未知"


def product_ids_filename_part(product_ids: list[str]) -> str:
    unique_ids = list(dict.fromkeys(product_ids))
    if not unique_ids:
        return "未知商品ID"
    if len(unique_ids) == 1:
        return unique_ids[0]
    preview = "+".join(unique_ids[:3])
    if len(unique_ids) > 3:
        preview += f"等{len(unique_ids)}个"
    return preview


def promotion_adjustment_update_date(period: str, fallback_date: dt.date) -> str:
    date_range = period_date_range(period)
    if date_range:
        return date_range[1]
    parsed = parse_date(period)
    return parsed.isoformat() if parsed else fallback_date.isoformat()


def promotion_adjustment_filename(shop_name: str, product_id_part: str, update_date: str) -> str:
    return "—".join(
        [
            sanitize_filename_part(shop_name),
            sanitize_filename_part(product_id_part),
            sanitize_filename_part(f"{PROMOTION_ADJUSTMENT_UPDATE_PREFIX}{update_date}"),
        ]
    )


def promotion_adjustment_existing_paths(target_dir: Path, shop_name: str, product_id_part: str) -> list[Path]:
    if not target_dir.exists():
        return []
    prefix = "—".join([sanitize_filename_part(shop_name), sanitize_filename_part(product_id_part), ""])
    return sorted(
        path
        for path in target_dir.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS and path.name.startswith(prefix)
    )


def after_sale_update_date(period: str, fallback_date: dt.date) -> str:
    date_range = period_date_range(period)
    if date_range:
        return date_range[1]
    parsed = parse_date(period)
    return parsed.isoformat() if parsed else fallback_date.isoformat()


def after_sale_filename(shop_name: str, update_date: str) -> str:
    return "—".join(
        [
            sanitize_filename_part(shop_name),
            sanitize_filename_part(f"{AFTER_SALE_UPDATE_PREFIX}{update_date}"),
        ]
    )


def after_sale_existing_paths(target_dir: Path) -> list[Path]:
    if not target_dir.exists():
        return []
    return sorted(
        path
        for path in target_dir.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )


def order_month_key_from_date(value: object) -> str:
    parsed = parse_datetime_text(value)
    return parsed.strftime("%Y-%m") if parsed else ""


def order_month_label(month_key: str) -> tuple[int, int]:
    year, month = month_key.split("-", 1)
    return int(year), int(month)


def order_data_folder(database_root: Path, shop_name: str, month_key: str) -> Path:
    year, month = order_month_label(month_key)
    return (
        database_root
        / ORDER_REPORT_TYPE
        / sanitize_filename_part(shop_name)
        / f"{year}年"
        / f"{month}月（成交时间）"
    )


def order_data_filename(shop_name: str, month_key: str) -> str:
    year, month = order_month_label(month_key)
    return "—".join(
        [
            sanitize_filename_part(shop_name),
            sanitize_filename_part(f"{ORDER_DATA_PREFIX}{year}年{month:02d}月"),
        ]
    )


def order_data_existing_paths(target_dir: Path) -> list[Path]:
    if not target_dir.exists():
        return []
    return sorted(
        path
        for path in target_dir.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS and not is_order_summary_file(path)
    )


def order_source_datetime(path: Path) -> dt.datetime:
    return download_datetime(path)


def is_cumulative_order_source_name(value: object) -> bool:
    source_name = normalize_text(value)
    if not source_name:
        return False
    return bool(
        re.search(
            rf"{re.escape(ORDER_DATA_PREFIX)}20\d{{2}}年\d{{1,2}}月(?:_第\d+部分)?",
            Path(source_name).name,
        )
    )


def saved_order_source_datetime(
    record: dict[str, str], cumulative_path: Path
) -> dt.datetime | None:
    saved_time = parse_datetime_text(record.get("汇总_最新下载时间", ""))
    saved_source = Path(normalize_text(record.get("汇总_最新来源文件", ""))).name
    saved_archive = Path(normalize_text(record.get("汇总_最新归档文件", ""))).name
    if saved_time is None or not saved_source:
        return None
    if (
        saved_source == cumulative_path.name
        or (saved_archive and saved_source == saved_archive)
        or is_cumulative_order_source_name(saved_source)
    ):
        return None
    return saved_time


def parse_datetime_text(value: object) -> dt.datetime | None:
    text = normalize_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            continue
    parsed_date = parse_date(text)
    return dt.datetime.combine(parsed_date, dt.time.min) if parsed_date else None


def is_ignorable_order_row(row: list[str]) -> bool:
    first_cell = normalize_text(row[0]) if row else ""
    return (
        first_cell in {"总计", "合计", "汇总"}
        or first_cell.startswith("注")
        or first_cell.startswith("说明")
    )


def resolve_order_row_datetime(
    headers: list[str], row: list[str]
) -> tuple[dt.datetime | None, str, str]:
    return resolve_order_row_datetime_from_columns(
        row, order_time_columns(headers)
    )


def resolve_order_row_datetime_from_columns(
    row: list[str], time_columns: list[tuple[str, int]]
) -> tuple[dt.datetime | None, str, str]:
    raw_values: list[str] = []
    for header, column_index in time_columns:
        raw_value = normalize_text(row[column_index]) if column_index < len(row) else ""
        raw_values.append(f"{header}={raw_value or '空'}")
        parsed = parse_datetime_text(raw_value)
        if parsed is not None:
            return parsed, header, "；".join(raw_values)
    return None, "", "；".join(raw_values) or "未找到订单时间列"


def scan_order_file_full(path: Path, max_cols: int) -> OrderFileScan:
    row_iterator = iter(iter_rows_full(path, max_cols))
    try:
        headers = next(row_iterator)
    except StopIteration as exc:
        raise RuntimeError(f"订单文件没有可读取内容：{path}") from exc

    order_index = find_header_index(headers, ORDER_ID_HEADERS)
    time_columns = order_time_columns(headers)
    if order_index is None:
        raise RuntimeError(f"订单文件缺少订单号列，停止整理并保留源文件：{path}")
    if not time_columns:
        raise RuntimeError(
            "订单文件缺少订单支付/付款时间、订单成交时间和订单创建时间列，"
            f"停止整理并保留源文件：{path}"
        )

    month_keys: set[str] = set()
    read_rows = 0
    valid_rows = 0
    ignored_rows = 0
    invalid_examples: list[str] = []
    invalid_count = 0
    for row_number, row in enumerate(row_iterator, start=2):
        if not any(normalize_text(cell) for cell in row):
            continue
        read_rows += 1
        if is_ignorable_order_row(row):
            ignored_rows += 1
            continue
        parsed_time, _selected_header, raw_values = resolve_order_row_datetime_from_columns(
            row, time_columns
        )
        if parsed_time is None:
            invalid_count += 1
            if len(invalid_examples) < 10:
                order_id = normalize_text(row[order_index]) if order_index < len(row) else ""
                invalid_examples.append(
                    f"行={row_number}，订单号={order_id or '空'}，时间字段={raw_values}"
                )
            continue
        valid_rows += 1
        month_keys.add(parsed_time.strftime("%Y-%m"))

    if invalid_count:
        examples = "；".join(invalid_examples)
        raise RuntimeError(
            f"订单时间无法识别，共{invalid_count}行；{examples}。"
            f"停止整理，不生成或覆盖月表，并保留源文件：{path}"
        )
    if read_rows != valid_rows + ignored_rows:
        raise RuntimeError(
            f"订单行数核对失败：读取={read_rows}，有效={valid_rows}，"
            f"忽略说明/合计={ignored_rows}；停止整理并保留源文件：{path}"
        )
    if not month_keys:
        raise RuntimeError(f"订单文件没有有效订单行，停止整理并保留源文件：{path}")
    return OrderFileScan(month_keys, read_rows, valid_rows, ignored_rows)


def order_merge_paths_full(
    analysis: Analysis, database_root: Path, max_cols: int
) -> tuple[list[Path], OrderFileScan]:
    source_scan = scan_order_file_full(analysis.source, max_cols)
    existing_paths: set[Path] = set()
    for month_key in source_scan.month_keys:
        target_dir = order_data_folder(database_root, analysis.shop_name, month_key)
        existing_paths.update(order_data_existing_paths(target_dir))
    merge_paths = sorted(existing_paths, key=lambda item: item.name)
    merge_paths.append(analysis.source)
    return merge_paths, source_scan


def validate_order_merge_inputs(
    analyses: list[Analysis], database_root: Path, max_cols: int
) -> None:
    checked: set[Path] = set()
    for analysis in analyses:
        if analysis.status != "ready" or analysis.report_type != ORDER_REPORT_TYPE:
            continue
        merge_paths, _source_scan = order_merge_paths_full(analysis, database_root, max_cols)
        for path in merge_paths:
            resolved = path.resolve()
            if resolved in checked:
                continue
            scan_order_file_full(path, max_cols)
            checked.add(resolved)
    if checked:
        print(f"order input integrity preflight passed: files={len(checked)}")


def create_order_merge_database(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(path)
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute(
        """
        CREATE TABLE order_records (
            month_key TEXT NOT NULL,
            order_id TEXT NOT NULL,
            sort_time TEXT NOT NULL,
            source_time TEXT NOT NULL,
            record_json TEXT NOT NULL,
            PRIMARY KEY (month_key, order_id)
        )
        """
    )
    return connection


def load_order_merge_database(
    connection: sqlite3.Connection,
    paths: list[Path],
    current_source: Path,
    shop_name: str,
    max_cols: int,
) -> tuple[dict[str, list[str]], int, int, int, int]:
    month_headers: dict[str, list[str]] = {}
    total_read = 0
    total_valid = 0
    total_ignored = 0
    duplicate_rows = 0
    pending_writes = 0

    for path in paths:
        row_iterator = iter(iter_rows_full(path, max_cols))
        try:
            headers = next(row_iterator)
        except StopIteration as exc:
            raise RuntimeError(f"订单文件没有可读取内容：{path}") from exc
        order_index = find_header_index(headers, ORDER_ID_HEADERS)
        time_columns = order_time_columns(headers)
        if order_index is None or not time_columns:
            raise RuntimeError(f"订单文件缺少订单号或可用时间列：{path}")
        is_current_source = path.resolve() == current_source.resolve()
        current_file_creation_time = (
            file_creation_datetime(path).strftime("%Y-%m-%d %H:%M:%S")
            if is_current_source
            else ""
        )
        path_read = 0
        path_valid = 0
        path_ignored = 0

        for row_number, row in enumerate(row_iterator, start=2):
            if not any(normalize_text(cell) for cell in row):
                continue
            path_read += 1
            if is_ignorable_order_row(row):
                path_ignored += 1
                continue
            parsed_time, _selected_header, raw_values = resolve_order_row_datetime_from_columns(
                row, time_columns
            )
            if parsed_time is None:
                order_id = normalize_text(row[order_index]) if order_index < len(row) else ""
                raise RuntimeError(
                    f"订单时间无法识别：文件={path}，行={row_number}，"
                    f"订单号={order_id or '空'}，时间字段={raw_values}；"
                    "停止整理并保留源文件。"
                )
            path_valid += 1
            month_key = parsed_time.strftime("%Y-%m")
            final_headers = month_headers.setdefault(month_key, [])
            append_headers(
                final_headers,
                [header for header in headers if header not in ORDER_SUMMARY_METADATA_HEADERS],
            )
            append_headers(final_headers, ORDER_SUMMARY_METADATA_HEADERS)
            order_id = normalize_text(row[order_index]) if order_index < len(row) else ""
            if not order_id:
                order_id = f"_row_{path.resolve()}_{row_number}"
            record = row_to_dict(headers, row)
            if is_current_source:
                source_time = current_file_creation_time
                record["汇总_最新下载时间"] = source_time
                record["汇总_最新来源文件"] = path.name
            else:
                saved_source_time = saved_order_source_datetime(record, path)
                source_time = (
                    saved_source_time.strftime("%Y-%m-%d %H:%M:%S")
                    if saved_source_time
                    else ""
                )
            record["汇总_店铺名称"] = shop_name
            record["汇总_最新归档文件"] = ""
            record_json = json.dumps(record, ensure_ascii=False, separators=(",", ":"))
            cursor = connection.execute(
                "INSERT OR IGNORE INTO order_records "
                "(month_key, order_id, sort_time, source_time, record_json) "
                "VALUES (?, ?, ?, ?, ?)",
                (month_key, order_id, parsed_time.isoformat(), source_time, record_json),
            )
            if cursor.rowcount == 0:
                duplicate_rows += 1
                old_row = connection.execute(
                    "SELECT sort_time, source_time, record_json FROM order_records "
                    "WHERE month_key=? AND order_id=?",
                    (month_key, order_id),
                ).fetchone()
                if old_row is None:
                    raise RuntimeError(
                        f"订单去重记录读取失败：月份={month_key}，订单号={order_id}"
                    )
                old_sort_time, old_source_time, old_record_json = old_row
                old_record = json.loads(old_record_json)
                if source_time >= old_source_time:
                    merged_record = merge_order_summary_record(
                        old_record, record, final_headers
                    )
                    latest_sort_time = parsed_time.isoformat()
                    latest_source_time = source_time
                else:
                    merged_record = merge_order_summary_record(
                        record, old_record, final_headers
                    )
                    latest_sort_time = old_sort_time
                    latest_source_time = old_source_time
                connection.execute(
                    "UPDATE order_records SET sort_time=?, source_time=?, record_json=? "
                    "WHERE month_key=? AND order_id=?",
                    (
                        latest_sort_time,
                        latest_source_time,
                        json.dumps(merged_record, ensure_ascii=False, separators=(",", ":")),
                        month_key,
                        order_id,
                    ),
                )
            pending_writes += 1
            if pending_writes >= STREAM_COMMIT_INTERVAL:
                connection.commit()
                pending_writes = 0

        if path_read != path_valid + path_ignored:
            raise RuntimeError(
                f"订单行数核对失败：文件={path}，读取={path_read}，"
                f"有效={path_valid}，忽略={path_ignored}"
            )
        total_read += path_read
        total_valid += path_valid
        total_ignored += path_ignored

    connection.commit()
    unique_rows = int(connection.execute("SELECT COUNT(*) FROM order_records").fetchone()[0])
    if total_valid - duplicate_rows != unique_rows:
        raise RuntimeError(
            f"订单去重核对失败：有效={total_valid}，重复={duplicate_rows}，唯一={unique_rows}"
        )
    return month_headers, total_read, total_valid, total_ignored, duplicate_rows


def order_output_part_paths(
    base_path: Path, total_rows: int, max_data_rows: int = EXCEL_MAX_DATA_ROWS
) -> list[Path]:
    if total_rows <= max_data_rows:
        return [base_path]
    part_count = math.ceil(total_rows / max_data_rows)
    width = max(2, len(str(part_count)))
    return [
        base_path.with_name(
            f"{base_path.stem}_第{part_number:0{width}d}部分{base_path.suffix}"
        )
        for part_number in range(1, part_count + 1)
    ]


def stage_order_outputs(
    connection: sqlite3.Connection,
    month_headers: dict[str, list[str]],
    shop_name: str,
    database_root: Path,
    staging_dir: Path,
    max_data_rows: int = EXCEL_MAX_DATA_ROWS,
) -> list[tuple[Path, Path, int]]:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - environment specific
        raise RuntimeError("openpyxl is required to write merged order data") from exc

    staging_dir.mkdir(parents=True, exist_ok=True)
    staged_outputs: list[tuple[Path, Path, int]] = []
    for month_key in sorted(month_headers):
        headers = month_headers[month_key]
        total_rows = int(
            connection.execute(
                "SELECT COUNT(*) FROM order_records WHERE month_key=?", (month_key,)
            ).fetchone()[0]
        )
        base_path = order_data_folder(database_root, shop_name, month_key) / (
            f"{order_data_filename(shop_name, month_key)}.xlsx"
        )
        final_paths = order_output_part_paths(base_path, total_rows, max_data_rows)
        records = connection.execute(
            "SELECT record_json FROM order_records WHERE month_key=? "
            "ORDER BY sort_time DESC, order_id",
            (month_key,),
        )
        remaining = total_rows
        month_written = 0
        for final_path in final_paths:
            part_rows = min(max_data_rows, remaining)
            staged_path = staging_dir / final_path.name
            workbook = Workbook(write_only=True)
            worksheet = workbook.create_sheet("推广调整日志")
            worksheet.freeze_panes = "A2"
            worksheet.append(headers)
            written = 0
            for _ in range(part_rows):
                database_row = records.fetchone()
                if database_row is None:
                    workbook.close()
                    raise RuntimeError(
                        f"订单输出提前结束：月份={month_key}，预计={total_rows}，已写={month_written}"
                    )
                record = json.loads(database_row[0])
                record["汇总_最新归档文件"] = final_path.name
                worksheet.append([normalize_text(record.get(header, "")) for header in headers])
                written += 1
                month_written += 1
            workbook.save(staged_path)
            workbook.close()
            staged_outputs.append((staged_path, final_path, written))
            remaining -= written
        if month_written != total_rows or remaining != 0:
            raise RuntimeError(
                f"订单输出行数核对失败：月份={month_key}，数据库={total_rows}，写入={month_written}"
            )
    return staged_outputs


def commit_order_outputs(
    staged_outputs: list[tuple[Path, Path, int]],
    old_paths: Iterable[Path],
    source_path: Path,
    backup_dir: Path,
) -> None:
    backup_dir.mkdir(parents=True, exist_ok=True)
    backups: list[tuple[Path, Path]] = []
    installed: list[Path] = []
    try:
        for index, old_path in enumerate(sorted(set(old_paths), key=lambda item: str(item))):
            if not old_path.exists():
                continue
            backup_path = backup_dir / f"{index:04d}_{old_path.name}"
            os.replace(old_path, backup_path)
            backups.append((backup_path, old_path))
        for staged_path, final_path, _row_count in staged_outputs:
            final_path.parent.mkdir(parents=True, exist_ok=True)
            os.replace(staged_path, final_path)
            installed.append(final_path)
    except Exception:
        for final_path in installed:
            final_path.unlink(missing_ok=True)
        for backup_path, old_path in reversed(backups):
            old_path.parent.mkdir(parents=True, exist_ok=True)
            os.replace(backup_path, old_path)
        raise
    source_path.unlink()


def promotion_adjustment_merge_rows(paths: list[Path], max_rows: int, max_cols: int) -> tuple[list[str], list[list[str]], str]:
    header: list[str] = []
    data_rows: list[list[str]] = []
    for path in paths:
        rows = read_rows(path, max_rows, max_cols)
        if not rows:
            continue
        if len(rows[0]) > len(header):
            header = rows[0]
        data_rows.extend(row for row in rows[1:] if any(normalize_text(cell) for cell in row))

    max_cols_seen = max([len(header), *(len(row) for row in data_rows)] or [0])
    if not header:
        header = ["操作时间", "操作人", "操作对象", "操作", "操作详情"]
    if len(header) < max_cols_seen:
        header = header + [f"列{index}" for index in range(len(header) + 1, max_cols_seen + 1)]

    time_index = find_header_index(header, ["操作时间"])
    latest_time: dt.datetime | None = None
    unique_rows: list[list[str]] = []
    seen: set[tuple[str, ...]] = set()
    for row in data_rows:
        normalized_row = [normalize_text(cell) for cell in row[:max_cols_seen]]
        if len(normalized_row) < max_cols_seen:
            normalized_row.extend([""] * (max_cols_seen - len(normalized_row)))
        key = tuple(normalized_row)
        if key in seen:
            continue
        seen.add(key)
        if time_index is not None and time_index < len(normalized_row):
            parsed_time = parse_datetime_text(normalized_row[time_index])
            if parsed_time and (latest_time is None or parsed_time > latest_time):
                latest_time = parsed_time
        unique_rows.append(normalized_row)

    if time_index is not None:
        unique_rows.sort(
            key=lambda row: (parse_datetime_text(row[time_index]) if time_index < len(row) else None) or dt.datetime.min,
            reverse=True,
        )
    latest_date = latest_time.date().isoformat() if latest_time else ""
    return header, unique_rows, latest_date


def merge_after_sale_rows(paths: list[Path], max_rows: int, max_cols: int) -> tuple[list[str], list[list[str]], str]:
    headers: list[str] = []
    rows_by_key: dict[tuple[str, str], list[str]] = {}
    fallback_index = 0
    latest_time: dt.datetime | None = None

    for path in paths:
        rows = read_rows(path, max_rows, max_cols)
        if not rows:
            continue
        append_headers(headers, rows[0])
        key_index = find_header_index(rows[0], ["售后编号", "退款编号"])
        time_index = find_header_index(rows[0], ["申请时间", "退款申请时间"])
        for row in rows[1:]:
            if not any(normalize_text(cell) for cell in row):
                continue
            record = {normalize_text(header): normalize_text(row[index]) if index < len(row) else "" for index, header in enumerate(rows[0])}
            append_headers(headers, record.keys())
            if key_index is not None and key_index < len(row):
                key_value = normalize_text(row[key_index])
                key = ("售后或退款编号", key_value) if key_value else ("row", "|".join(normalize_text(cell) for cell in row))
            else:
                key = ("row", "|".join(normalize_text(cell) for cell in row))
            if not key[1]:
                fallback_index += 1
                key = ("empty", f"{fallback_index}")
            output_row = [record.get(header, "") for header in headers]
            rows_by_key[key] = output_row
            if time_index is not None and time_index < len(row):
                parsed_time = parse_datetime_text(row[time_index])
                if parsed_time and (latest_time is None or parsed_time > latest_time):
                    latest_time = parsed_time

    normalized_rows: list[list[str]] = []
    for row in rows_by_key.values():
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        normalized_rows.append(row[: len(headers)])

    sort_index = find_header_index(headers, ["申请时间", "退款申请时间"])
    if sort_index is not None:
        normalized_rows.sort(
            key=lambda row: (parse_datetime_text(row[sort_index]) if sort_index < len(row) else None) or dt.datetime.min,
            reverse=True,
        )
    latest_date = latest_time.date().isoformat() if latest_time else ""
    return headers, normalized_rows, latest_date


def write_xlsx_table(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - environment specific
        raise RuntimeError("openpyxl is required to write merged promotion adjustment logs") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "推广调整日志"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(normalize_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)

    temp_path = path.with_name(f"{path.stem}.tmp{path.suffix}")
    workbook.save(temp_path)
    workbook.close()
    os.replace(temp_path, path)


def merge_promotion_adjustment_log(analysis: Analysis, max_rows: int, max_cols: int) -> tuple[Path, int, int]:
    target_dir = analysis.target_path.parent
    product_id_part = product_ids_filename_part(analysis.product_ids)
    existing_paths = promotion_adjustment_existing_paths(target_dir, analysis.shop_name, product_id_part)
    merge_paths = [*existing_paths, analysis.source]
    headers, rows, latest_date = promotion_adjustment_merge_rows(merge_paths, max_rows, max_cols)
    if not latest_date:
        latest_date = promotion_adjustment_update_date(analysis.period, dt.date.fromisoformat(analysis.download_date))
    final_name = promotion_adjustment_filename(analysis.shop_name, product_id_part, latest_date)
    final_path = target_dir / f"{final_name}.xlsx"
    write_xlsx_table(final_path, headers, rows)
    for path in merge_paths:
        if path.resolve() == final_path.resolve():
            continue
        if path.exists():
            path.unlink()
    analysis.target_path = final_path
    return final_path, len(merge_paths), len(rows)


def merge_after_sale_data(analysis: Analysis, max_rows: int, max_cols: int) -> tuple[Path, int, int]:
    target_dir = analysis.target_path.parent
    existing_paths = after_sale_existing_paths(target_dir)
    merge_paths = [*existing_paths, analysis.source]
    headers, rows, latest_date = merge_after_sale_rows(merge_paths, max_rows, max_cols)
    if not latest_date:
        latest_date = after_sale_update_date(analysis.period, dt.date.fromisoformat(analysis.download_date))
    final_path = target_dir / f"{after_sale_filename(analysis.shop_name, latest_date)}.xlsx"
    write_xlsx_table(final_path, headers, rows)
    for path in merge_paths:
        if path.resolve() == final_path.resolve():
            continue
        if path.exists():
            path.unlink()
    analysis.target_path = final_path
    return final_path, len(merge_paths), len(rows)


def order_month_keys_from_rows(rows: list[list[str]]) -> set[str]:
    if not rows:
        return set()
    headers = rows[0]
    time_index = find_header_index(headers, ORDER_TIME_HEADERS)
    if time_index is None:
        return set()
    month_keys: set[str] = set()
    for row in rows[1:]:
        if time_index < len(row):
            month_key = order_month_key_from_date(row[time_index])
            if month_key:
                month_keys.add(month_key)
    return month_keys


def merge_order_rows_by_month(
    paths: list[Path],
    shop_name: str,
    max_rows: int,
    max_cols: int,
) -> dict[str, tuple[list[str], list[list[str]]]]:
    month_headers: dict[str, list[str]] = {}
    month_records: dict[str, dict[str, dict[str, str]]] = {}
    fallback_index = 0

    for path in sorted(paths, key=lambda item: (order_source_datetime(item), item.name)):
        rows = read_rows(path, max_rows, max_cols)
        if not rows:
            continue
        headers = rows[0]
        order_index = find_header_index(headers, ORDER_ID_HEADERS)
        time_index = find_header_index(headers, ORDER_TIME_HEADERS)
        if order_index is None or time_index is None:
            continue
        source_time = order_source_datetime(path).strftime("%Y-%m-%d %H:%M:%S")
        for row in rows[1:]:
            if not any(normalize_text(cell) for cell in row):
                continue
            if time_index >= len(row):
                continue
            month_key = order_month_key_from_date(row[time_index])
            if not month_key:
                continue
            final_headers = month_headers.setdefault(month_key, [])
            append_headers(final_headers, [header for header in headers if header not in ORDER_SUMMARY_METADATA_HEADERS])
            append_headers(final_headers, ORDER_SUMMARY_METADATA_HEADERS)
            records = month_records.setdefault(month_key, {})
            order_id = normalize_text(row[order_index]) if order_index < len(row) else ""
            if not order_id:
                fallback_index += 1
                order_id = f"_row_{fallback_index}"
            record = row_to_dict(headers, row)
            record["汇总_店铺名称"] = shop_name
            record["汇总_最新下载时间"] = source_time
            record["汇总_最新来源文件"] = path.name
            record["汇总_最新归档文件"] = ""
            if order_id in records:
                records[order_id] = merge_order_summary_record(records[order_id], record, final_headers)
            else:
                records[order_id] = {header: normalize_text(record.get(header, "")) for header in final_headers}

    result: dict[str, tuple[list[str], list[list[str]]]] = {}
    for month_key, records in month_records.items():
        headers = month_headers[month_key]
        time_index = find_header_index(headers, ORDER_TIME_HEADERS)
        rows = []
        for record in records.values():
            rows.append([normalize_text(record.get(header, "")) for header in headers])
        if time_index is not None:
            rows.sort(
                key=lambda row: (parse_datetime_text(row[time_index]) if time_index < len(row) else None) or dt.datetime.min,
                reverse=True,
            )
        result[month_key] = (headers, rows)
    return result


def merge_order_data(
    analysis: Analysis,
    database_root: Path,
    max_rows: int,
    max_cols: int,
) -> list[tuple[Path, int, int]]:
    del max_rows  # The row cap is preview-only; formal order merging always reads every row.
    merge_paths, source_scan = order_merge_paths_full(analysis, database_root, max_cols)
    old_paths = [path for path in merge_paths if path.resolve() != analysis.source.resolve()]
    with tempfile.TemporaryDirectory(prefix=".order-merge-", dir=database_root) as temp_name:
        transaction_root = Path(temp_name)
        connection = create_order_merge_database(transaction_root / "orders.sqlite3")
        try:
            (
                month_headers,
                total_read,
                total_valid,
                total_ignored,
                duplicate_rows,
            ) = load_order_merge_database(
                connection,
                merge_paths,
                analysis.source,
                analysis.shop_name,
                max_cols,
            )
            staged_outputs = stage_order_outputs(
                connection,
                month_headers,
                analysis.shop_name,
                database_root,
                transaction_root / "new",
            )
            written_rows = sum(row_count for _staged, _final, row_count in staged_outputs)
            unique_rows = int(
                connection.execute("SELECT COUNT(*) FROM order_records").fetchone()[0]
            )
            if written_rows != unique_rows:
                raise RuntimeError(
                    f"订单合并最终核对失败：唯一订单={unique_rows}，写入={written_rows}；"
                    "停止整理并保留源文件。"
                )
            commit_order_outputs(
                staged_outputs,
                old_paths,
                analysis.source,
                transaction_root / "old",
            )
        finally:
            connection.close()

    outputs = [
        (final_path, len(merge_paths), row_count)
        for _staged_path, final_path, row_count in staged_outputs
    ]
    if outputs:
        analysis.target_path = outputs[-1][0]
    print(
        "order row integrity verified:",
        f"source_rows={source_scan.read_rows}",
        f"all_read={total_read}",
        f"valid={total_valid}",
        f"ignored={total_ignored}",
        f"duplicates={duplicate_rows}",
        f"written={sum(row_count for _path, _sources, row_count in outputs)}",
    )
    return outputs


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    index = 2
    while True:
        candidate = parent / f"{stem}_{index:02d}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def reserve_unique_targets(analyses: list[Analysis]) -> None:
    reserved: set[str] = set()
    for analysis in analyses:
        if analysis.status != "ready":
            continue
        if analysis.report_type in {ORDER_REPORT_TYPE, AFTER_SALE_REPORT_TYPE, PROMOTION_ADJUSTMENT_REPORT_TYPE}:
            continue
        original = analysis.target_path
        candidate = original
        index = 2
        while str(candidate).lower() in reserved or candidate.exists():
            candidate = original.parent / f"{original.stem}_{index:02d}{original.suffix}"
            index += 1
        analysis.target_path = candidate
        reserved.add(str(candidate).lower())


def target_for_pending(database_root: Path, source: Path, file_download_date: dt.date) -> Path:
    return source


def target_for_success(
    database_root: Path,
    source: Path,
    signature: Signature,
    shop_name: str,
    period: str,
    file_download_date: dt.date,
    product_ids: list[str] | None = None,
) -> Path:
    target_dir = (
        database_root
        / signature.target_folder
        / sanitize_filename_part(shop_name)
        / f"{file_download_date.year}年"
        / f"{file_download_date.month}月（导出时间）"
    )
    if signature.report_type == PROMOTION_ADJUSTMENT_REPORT_TYPE:
        update_date = promotion_adjustment_update_date(period, file_download_date)
        filename = promotion_adjustment_filename(shop_name, product_ids_filename_part(product_ids or []), update_date)
        return target_dir / f"{filename}.xlsx"
    if signature.report_type == ORDER_REPORT_TYPE:
        date_range = period_date_range(period)
        month_key = (date_range[1] if date_range else file_download_date.isoformat())[:7]
        target_dir = order_data_folder(database_root, shop_name, month_key)
        return target_dir / f"{order_data_filename(shop_name, month_key)}.xlsx"
    if signature.report_type == AFTER_SALE_REPORT_TYPE:
        update_date = after_sale_update_date(period, file_download_date)
        filename = after_sale_filename(shop_name, update_date)
        return target_dir / f"{filename}.xlsx"
    else:
        filename = "-".join(
            [
                file_download_date.isoformat(),
                sanitize_filename_part(shop_name),
                sanitize_filename_part(signature.report_type),
                sanitize_filename_part(period),
            ]
        )
    return unique_path(target_dir / f"{filename}{source.suffix.lower()}")


def analyze_file(
    source: Path,
    database_root: Path,
    signatures: list[Signature],
    product_map: dict[str, str],
    shops: list[str],
    max_rows: int,
    max_cols: int,
) -> Analysis:
    file_download_date = download_date(source)
    try:
        rows = read_rows(source, max_rows, max_cols)
    except Exception as exc:
        target = target_for_pending(database_root, source, file_download_date)
        return Analysis(
            source=source,
            status="pending",
            reason=f"read_error: {exc}",
            report_type="",
            target_folder="_待确认",
            shop_name="",
            shop_source="",
            period="",
            period_source="",
            download_date=file_download_date.isoformat(),
            target_path=target,
            product_counts="",
            matched_headers="",
            missing_fields="read_error",
            map_updates=[],
            product_ids=[],
            new_product_ids=[],
        )

    signature, header_row_index, headers, matched = match_signature(rows, signatures)
    if signature is None:
        target = target_for_pending(database_root, source, file_download_date)
        return Analysis(
            source=source,
            status="pending",
            reason="unknown_report_signature",
            report_type="",
            target_folder="_待确认",
            shop_name="",
            shop_source="",
            period="",
            period_source="",
            download_date=file_download_date.isoformat(),
            target_path=target,
            product_counts="",
            matched_headers="",
            missing_fields="report_type",
            map_updates=[],
            product_ids=[],
            new_product_ids=[],
        )

    product_ids = collect_product_ids(rows, header_row_index, headers, signature)
    new_product_ids = []
    if signature.report_type in PRODUCT_ID_CHECK_REPORT_TYPES:
        new_product_ids = sorted({product_id for product_id in product_ids if product_id not in product_map})
    shop_name, shop_source, product_counts = decide_shop(
        source, signature.report_type, rows, product_ids, product_map, shops
    )
    platform, platform_source = detect_platform(source, rows, signature)
    shop_name = normalize_shop_platform(shop_name, platform)
    if shop_name and platform_source:
        shop_source = "+".join(part for part in (shop_source, f"platform:{platform_source}") if part)
    period, period_source = decide_period(source, signature, rows, header_row_index, headers, file_download_date)

    missing: list[str] = []
    if not shop_name:
        missing.append("shop_name")
    if not period:
        missing.append("period")
    if is_multi_day_promotion_summary(
        source,
        signature.report_type,
        period,
        rows,
        header_row_index,
        headers,
        signature.period_headers,
        signature.product_id_headers,
    ):
        missing.append("single_day_period")

    map_updates: list[tuple[str, str]] = []
    if signature.report_type == "商品数据" and shop_name:
        for product_id in product_ids:
            if product_id and product_map.get(product_id) != shop_name:
                map_updates.append((product_id, shop_name))

    if missing:
        target = target_for_pending(database_root, source, file_download_date)
        reason = "multi_day_promotion_summary" if missing == ["single_day_period"] else "missing_required_fields"
        return Analysis(
            source=source,
            status="pending",
            reason=reason,
            report_type=signature.report_type,
            target_folder="_待确认",
            shop_name=shop_name,
            shop_source=shop_source,
            period=period,
            period_source=period_source,
            download_date=file_download_date.isoformat(),
            target_path=target,
            product_counts=product_counts,
            matched_headers="|".join(matched),
            missing_fields="|".join(missing),
            map_updates=map_updates,
            product_ids=product_ids,
            new_product_ids=new_product_ids,
        )

    target = target_for_success(database_root, source, signature, shop_name, period, file_download_date, product_ids)
    return Analysis(
        source=source,
        status="ready",
        reason="ok",
        report_type=signature.report_type,
        target_folder=signature.target_folder,
        shop_name=shop_name,
        shop_source=shop_source,
        period=period,
        period_source=period_source,
        download_date=file_download_date.isoformat(),
        target_path=target,
        product_counts=product_counts,
        matched_headers="|".join(matched),
        missing_fields="",
        map_updates=map_updates,
        product_ids=product_ids,
        new_product_ids=new_product_ids,
    )


def iter_input_files(input_dir: Path) -> list[Path]:
    return sorted(
        [path for path in input_dir.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS],
        key=lambda path: path.name.lower(),
    )


def file_crc32(path: Path) -> int:
    checksum = 0
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            checksum = zlib.crc32(chunk, checksum)
    return checksum & 0xFFFFFFFF


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def mark_exact_duplicate_inputs(analyses: list[Analysis], input_dir: Path) -> None:
    """Keep one deterministically chosen source per identical-content group."""
    groups: dict[tuple[int, str], list[Analysis]] = {}
    for analysis in analyses:
        if not analysis.source.is_file():
            continue
        key = (analysis.source.stat().st_size, file_sha256(analysis.source))
        groups.setdefault(key, []).append(analysis)

    duplicate_dir = input_dir / DUPLICATE_INPUT_FOLDER_NAME
    for group in groups.values():
        if len(group) < 2:
            continue
        group.sort(key=lambda item: (item.status != "ready", item.source.name.lower()))
        retained = group[0]
        for duplicate in group[1:]:
            duplicate.status = "duplicate_file"
            duplicate.reason = f"exact_duplicate_of:{retained.source.name}"
            duplicate.target_folder = DUPLICATE_INPUT_FOLDER_NAME
            duplicate.target_path = unique_path(duplicate_dir / duplicate.source.name)


def zip_member_filename(info: zipfile.ZipInfo) -> str:
    raw_name = info.filename.replace("\\", "/")
    member_path = PurePosixPath(raw_name)
    if (
        member_path.is_absolute()
        or ".." in member_path.parts
        or re.match(r"^[A-Za-z]:", raw_name)
    ):
        raise RuntimeError(f"unsafe ZIP member path: {info.filename}")
    filename = member_path.name
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        return ""
    stem = sanitize_filename_part(Path(filename).stem) or "report"
    return f"{stem}{suffix}"


def unique_extraction_target(input_dir: Path, filename: str, reserved: set[Path]) -> Path:
    candidate = input_dir / filename
    if candidate not in reserved and not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    index = 2
    while True:
        alternative = input_dir / f"{stem}_{index:02d}{suffix}"
        if alternative not in reserved and not alternative.exists():
            return alternative
        index += 1


def extract_zip_inputs(input_dir: Path) -> None:
    zip_paths = sorted(
        [
            path
            for path in input_dir.iterdir()
            if path.is_file() and path.suffix.lower() in SUPPORTED_ARCHIVE_EXTENSIONS
        ],
        key=lambda path: path.name.lower(),
    )
    if not zip_paths:
        return

    extracted_zip_dir = input_dir / EXTRACTED_ZIP_FOLDER_NAME
    for zip_path in zip_paths:
        temp_paths: list[Path] = []
        committed_paths: list[Path] = []
        try:
            with zipfile.ZipFile(zip_path) as archive:
                members: list[tuple[zipfile.ZipInfo, str]] = []
                for info in archive.infolist():
                    if info.is_dir():
                        continue
                    filename = zip_member_filename(info)
                    if filename:
                        members.append((info, filename))

                if not members:
                    raise RuntimeError("no supported .csv, .xlsx, or .xls files found")

                reserved: set[Path] = set()
                planned: list[tuple[Path, Path]] = []
                reused_count = 0
                for info, filename in members:
                    preferred = input_dir / filename
                    if (
                        preferred not in reserved
                        and preferred.exists()
                        and preferred.is_file()
                        and preferred.stat().st_size == info.file_size
                        and file_crc32(preferred) == info.CRC
                    ):
                        reserved.add(preferred)
                        reused_count += 1
                        continue

                    target = unique_extraction_target(input_dir, filename, reserved)
                    reserved.add(target)
                    with tempfile.NamedTemporaryFile(
                        prefix=".zip_extract_", suffix=".tmp", dir=input_dir, delete=False
                    ) as temp_handle:
                        temp_path = Path(temp_handle.name)
                    temp_paths.append(temp_path)
                    with archive.open(info) as source, temp_path.open("wb") as destination:
                        shutil.copyfileobj(source, destination)
                    planned.append((temp_path, target))

                for temp_path, target in planned:
                    os.replace(temp_path, target)
                    committed_paths.append(target)
                    temp_paths.remove(temp_path)

            extracted_zip_dir.mkdir(parents=True, exist_ok=True)
            archived_zip = unique_path(extracted_zip_dir / zip_path.name)
            shutil.move(str(zip_path), str(archived_zip))
            print(
                "zip extracted:",
                f"source={zip_path}",
                f"new_files={len(committed_paths)}",
                f"existing_files={reused_count}",
                f"zip_archived={archived_zip}",
            )
        except (OSError, RuntimeError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
            for temp_path in temp_paths:
                try:
                    temp_path.unlink(missing_ok=True)
                except OSError:
                    pass
            for committed_path in committed_paths:
                try:
                    committed_path.unlink(missing_ok=True)
                except OSError:
                    pass
            raise RuntimeError(f"ZIP preparation failed for {zip_path}: {exc}") from exc


def input_file_prefix(path: Path) -> str:
    match = re.match(r"([0-9a-f]{32})", path.name, flags=re.IGNORECASE)
    return match.group(1).lower() if match else ""


def apply_batch_shop_inference(
    database_root: Path,
    analyses: list[Analysis],
    signatures: list[Signature],
    file_download_dates: dict[Path, dt.date],
) -> None:
    signatures_by_type = {signature.report_type: signature for signature in signatures}
    prefix_shops: dict[str, str] = {}
    for analysis in analyses:
        prefix = input_file_prefix(analysis.source)
        if prefix and analysis.shop_name:
            prefix_shops.setdefault(prefix, analysis.shop_name)

    batch_product_shops: dict[str, str] = {}
    for analysis in analyses:
        if not analysis.shop_name:
            continue
        for product_id in analysis.new_product_ids:
            batch_product_shops.setdefault(product_id, analysis.shop_name)

    for analysis in analyses:
        if analysis.shop_name or not analysis.report_type or not analysis.period:
            continue
        prefix = input_file_prefix(analysis.source)
        shop_name = prefix_shops.get(prefix, "")
        if not shop_name and analysis.new_product_ids:
            counts = Counter(
                batch_product_shops[product_id]
                for product_id in analysis.new_product_ids
                if product_id in batch_product_shops
            )
            if counts:
                top_shop, top_count = counts.most_common(1)[0]
                second_count = counts.most_common(2)[1][1] if len(counts) > 1 else 0
                if top_count > second_count:
                    shop_name = top_shop
        signature = signatures_by_type.get(analysis.report_type)
        if not shop_name or signature is None:
            continue
        analysis.shop_name = shop_name
        analysis.shop_source = "batch_prefix" if prefix and prefix in prefix_shops else "batch_product_id"
        analysis.missing_fields = ";".join(
            field for field in analysis.missing_fields.split(";") if field and field != "shop_name"
        )
        if not analysis.missing_fields:
            analysis.status = "ready"
            analysis.reason = "ok"
            analysis.target_folder = signature.target_folder
            analysis.target_path = target_for_success(
                database_root,
                analysis.source,
                signature,
                shop_name,
                analysis.period,
                file_download_dates.get(analysis.source, download_date(analysis.source)),
                analysis.product_ids,
            )


def manifest_row(analysis: Analysis, action: str) -> dict[str, str]:
    return {
        "action": action,
        "status": analysis.status,
        "reason": analysis.reason,
        "source_path": str(analysis.source),
        "source_name": analysis.source.name,
        "target_path": str(analysis.target_path),
        "target_name": analysis.target_path.name,
        "report_type": analysis.report_type,
        "target_folder": analysis.target_folder,
        "shop_name": analysis.shop_name,
        "shop_source": analysis.shop_source,
        "period": analysis.period,
        "period_source": analysis.period_source,
        "download_date": analysis.download_date,
        "product_counts": analysis.product_counts,
        "matched_headers": analysis.matched_headers,
        "missing_fields": analysis.missing_fields,
        "map_updates": ";".join(f"{product_id}:{shop}" for product_id, shop in analysis.map_updates),
        "new_product_ids": "|".join(analysis.new_product_ids),
    }


def missing_reminder_path(database_root: Path, apply: bool) -> Path:
    return database_root / ("缺失提醒.csv" if apply else "缺失提醒_预览.csv")


def manifest_output_path(database_root: Path, apply: bool, apply_blocked: bool) -> Path:
    if apply_blocked:
        return database_root / "整理记录_中断.csv"
    return database_root / ("整理记录.csv" if apply else "整理记录_预览.csv")


def has_missing_required_reports(reminder_rows: list[dict[str, str]]) -> bool:
    return any(row["status"] == "missing" for row in reminder_rows)


def is_pending_input_dir(input_dir: Path) -> bool:
    return "_待确认" in input_dir.parts


def missing_report_reminder_rows(
    expected_reports: list[ExpectedReport],
    analyses: list[Analysis],
    action: str,
    input_dir: Path,
) -> list[dict[str, str]]:
    detected_counts = Counter(
        analysis.report_type
        for analysis in analyses
        if normalize_text(analysis.report_type)
    )
    rows: list[dict[str, str]] = []
    if not is_pending_input_dir(input_dir):
        for expected in expected_reports:
            detected_count = detected_counts.get(expected.report_type, 0)
            if expected.required and detected_count == 0:
                message = expected.reminder_message or f"本次未检测到必备报表：{expected.report_type}"
                rows.append(
                    {
                        "action": action,
                        "status": "missing",
                        "report_type": expected.report_type,
                        "detected_count": "0",
                        "shop_name": "",
                        "source_name": "",
                        "source_path": "",
                        "new_product_ids": "",
                        "message": message,
                        "input_path": str(input_dir),
                    }
                )

    def preview_product_ids(product_ids: list[str]) -> str:
        preview_ids = "、".join(product_ids[:10])
        if len(product_ids) > 10:
            preview_ids += f" 等{len(product_ids)}个"
        return preview_ids

    product_id_groups: dict[str, dict[str, dict[str, set[str]]]] = {}
    for analysis in analyses:
        if analysis.status == "duplicate_file":
            continue
        if analysis.report_type not in PRODUCT_ID_CHECK_REPORT_TYPES or not analysis.new_product_ids:
            continue
        shop_key = analysis.shop_name or "未识别店铺"
        shop_group = product_id_groups.setdefault(shop_key, {})
        for product_id in analysis.new_product_ids:
            product_group = shop_group.setdefault(
                product_id,
                {"report_types": set(), "source_names": set(), "source_paths": set()},
            )
            product_group["report_types"].add(analysis.report_type)
            product_group["source_names"].add(analysis.source.name)
            product_group["source_paths"].add(str(analysis.source))

    for shop_name in sorted(product_id_groups):
        group = product_id_groups[shop_name]
        product_ids = sorted(
            product_id
            for product_id, product_group in group.items()
            if product_group["report_types"] & PRODUCT_DATA_REQUIRED_SOURCE_REPORT_TYPES
        )
        if product_ids:
            report_types = sorted(
                {
                    report_type
                    for product_id in product_ids
                    for report_type in group[product_id]["report_types"]
                    if report_type in PRODUCT_DATA_REQUIRED_SOURCE_REPORT_TYPES
                }
            )
            source_names = sorted(
                {
                    source_name
                    for product_id in product_ids
                    for source_name in group[product_id]["source_names"]
                }
            )
            source_paths = sorted(
                {
                    source_path
                    for product_id in product_ids
                    for source_path in group[product_id]["source_paths"]
                }
            )
            message = (
                f"{shop_name} 的{'/'.join(report_types)}中出现商品数据未收录的商品ID：{preview_product_ids(product_ids)}；"
                "请更新或补导该店铺的商品数据。"
            )
            rows.append(
                {
                    "action": action,
                    "status": "new_product_id",
                    "report_type": "商品数据",
                    "detected_count": str(len(product_ids)),
                    "shop_name": "" if shop_name == "未识别店铺" else shop_name,
                    "source_name": ";".join(source_names),
                    "source_path": ";".join(source_paths),
                    "new_product_ids": "|".join(product_ids),
                    "message": message,
                    "input_path": str(input_dir),
                }
            )

        historical_product_ids = sorted(
            product_id
            for product_id, product_group in group.items()
            if product_group["report_types"] and product_group["report_types"].issubset(PRODUCT_HISTORY_SOURCE_REPORT_TYPES)
        )
        if not historical_product_ids:
            continue
        source_names = sorted(
            {
                source_name
                for product_id in historical_product_ids
                for source_name in group[product_id]["source_names"]
            }
        )
        source_paths = sorted(
            {
                source_path
                for product_id in historical_product_ids
                for source_path in group[product_id]["source_paths"]
            }
        )
        message = (
            f"{shop_name} 的售后数据中出现商品数据未收录的历史/疑似已下架商品ID："
            f"{preview_product_ids(historical_product_ids)}；这类ID仅用于售后记录核对，不要求补导商品数据。"
        )
        rows.append(
            {
                "action": action,
                "status": "historical_product_id",
                "report_type": "售后数据",
                "detected_count": str(len(historical_product_ids)),
                "shop_name": "" if shop_name == "未识别店铺" else shop_name,
                "source_name": ";".join(source_names),
                "source_path": ";".join(source_paths),
                "new_product_ids": "|".join(historical_product_ids),
                "message": message,
                "input_path": str(input_dir),
            }
        )

    def pending_reminder_message(analysis: Analysis) -> str:
        if analysis.reason == "multi_day_promotion_summary":
            shop_name = analysis.shop_name or "未识别店铺"
            period_text = f"（{analysis.period}）" if analysis.period else ""
            return (
                f"{shop_name} 的推广数据{period_text}是多日推广汇总，不归档；"
                "文件已保留在原位置，可能需要按单日重新下载该店铺推广数据。"
            )
        return f"文件未整理，保留在原位置：{analysis.reason}；缺少字段：{analysis.missing_fields}"

    for analysis in analyses:
        if analysis.status != "duplicate_file":
            continue
        retained_name = analysis.reason.removeprefix("exact_duplicate_of:")
        rows.append(
            {
                "action": action,
                "status": "duplicate_file",
                "report_type": analysis.report_type,
                "detected_count": "1",
                "shop_name": analysis.shop_name,
                "source_name": analysis.source.name,
                "source_path": str(analysis.source),
                "new_product_ids": "",
                "message": (
                    f"文件内容与 {retained_name} 完全相同；仅保留该文件参与归档/合并，"
                    f"本文件在正式处理时将移至 {DUPLICATE_INPUT_FOLDER_NAME}。"
                ),
                "input_path": str(input_dir),
            }
        )

    for analysis in analyses:
        if analysis.status != "pending":
            continue
        rows.append(
            {
                "action": action,
                "status": "pending",
                "report_type": analysis.report_type,
                "detected_count": "",
                "shop_name": analysis.shop_name,
                "source_name": analysis.source.name,
                "source_path": str(analysis.source),
                "new_product_ids": "|".join(analysis.new_product_ids),
                "message": pending_reminder_message(analysis),
                "input_path": str(input_dir),
            }
        )

    if not rows:
        rows.append(
            {
                "action": action,
                "status": "ok",
                "report_type": "",
                "detected_count": "",
                "shop_name": "",
                "source_name": "",
                "source_path": "",
                "new_product_ids": "",
                "message": "本次未发现必备报表缺失、需补导商品数据的新商品ID或售后历史商品ID。",
                "input_path": str(input_dir),
            }
        )
    return rows


def print_missing_reminders(rows: list[dict[str, str]], reminder_path: Path) -> None:
    print(f"reminder log: {reminder_path}")
    print("完整提醒日志已写入提醒表，并同步显示在当前消息窗口：")
    for index, row in enumerate(rows, start=1):
        status = row.get("status", "") or "unknown"
        shop_name = row.get("shop_name", "") or "未指定"
        report_type = row.get("report_type", "") or "未指定"
        source_name = row.get("source_name", "") or "未指定"
        product_ids = row.get("new_product_ids", "") or "无"
        print(
            f"提醒日志 {index}/{len(rows)}：状态={status}；店铺={shop_name}；"
            f"报表={report_type}；来源文件={source_name}；商品ID={product_ids}"
        )
        print(f"  内容：{row.get('message', '')}")


def order_summary_month_key(value: dt.datetime) -> str:
    return value.strftime("%Y-%m")


def order_summary_name(month_key: str) -> str:
    year, month = month_key.split("-", 1)
    return f"{ORDER_SUMMARY_PREFIX}_{year}年{month}月.csv"


def order_summary_path(database_root: Path, month_key: str) -> Path:
    return database_root / "订单数据" / order_summary_name(month_key)


def is_order_summary_file(path: Path) -> bool:
    return path.suffix.lower() == ".csv" and path.stem.startswith(ORDER_SUMMARY_PREFIX)


def analysis_summary_datetime(analysis: Analysis) -> dt.datetime:
    if analysis.target_path.exists():
        return download_datetime(analysis.target_path)
    return download_datetime(analysis.source)


def analysis_summary_month_key(analysis: Analysis) -> str:
    return order_summary_month_key(analysis_summary_datetime(analysis))


def order_summary_preview_paths(database_root: Path, analyses: list[Analysis]) -> list[Path]:
    month_keys = sorted(
        {
            analysis_summary_month_key(analysis)
            for analysis in analyses
            if analysis.status == "ready" and analysis.report_type == "订单数据"
        }
    )
    return [order_summary_path(database_root, month_key) for month_key in month_keys]


def read_existing_order_summary(path: Path) -> tuple[list[str], dict[str, dict[str, str]]]:
    if not path.exists():
        return [], {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        rows: dict[str, dict[str, str]] = {}
        order_key = find_header_index(headers, ORDER_ID_HEADERS)
        if order_key is None:
            return headers, {}
        order_header = headers[order_key]
        for row in reader:
            order_id = normalize_text(row.get(order_header, ""))
            if order_id:
                rows[order_id] = {header: normalize_text(row.get(header, "")) for header in headers}
        return headers, rows


def row_to_dict(headers: list[str], row: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for index, header in enumerate(headers):
        clean_header = normalize_text(header)
        if not clean_header:
            continue
        result[clean_header] = normalize_text(row[index]) if index < len(row) else ""
    return result


def is_protected_order_info_header(header: str) -> bool:
    clean_header = normalize_text(header)
    if clean_header in PROTECTED_ORDER_INFO_EXACT_HEADERS:
        return True
    if any(keyword in clean_header for keyword in PROTECTED_ORDER_INFO_EXCLUDE_KEYWORDS):
        return False
    return any(keyword in clean_header for keyword in PROTECTED_ORDER_INFO_KEYWORDS)


def is_masked_or_empty(value: str) -> bool:
    text = normalize_text(value)
    if not text:
        return True
    return "*" in text or "＊" in text or "加密" in text


def merge_order_summary_record(
    old_record: dict[str, str],
    new_record: dict[str, str],
    final_headers: list[str],
) -> dict[str, str]:
    merged: dict[str, str] = {}
    for header in final_headers:
        old_value = normalize_text(old_record.get(header, ""))
        if header not in new_record:
            merged[header] = old_value
            continue
        new_value = normalize_text(new_record.get(header, ""))
        if not new_value and old_value:
            merged[header] = old_value
            continue
        if (
            is_protected_order_info_header(header)
            and is_masked_or_empty(new_value)
            and not is_masked_or_empty(old_value)
        ):
            merged[header] = old_value
        else:
            merged[header] = new_value
    return merged


def append_headers(target: list[str], headers: Iterable[str]) -> None:
    existing = set(target)
    for header in headers:
        clean_header = normalize_text(header)
        if clean_header and clean_header not in existing:
            target.append(clean_header)
            existing.add(clean_header)


def update_order_summary(
    database_root: Path,
    analyses: list[Analysis],
    signatures: list[Signature],
    max_rows: int,
    max_cols: int,
    reset: bool = False,
) -> list[tuple[Path, int, int, int, int]]:
    order_analyses = [
        analysis
        for analysis in analyses
        if analysis.status == "ready" and analysis.report_type == "订单数据" and analysis.target_path.exists()
    ]
    if not order_analyses:
        return []

    grouped_analyses: dict[str, list[Analysis]] = {}
    for analysis in order_analyses:
        month_key = analysis_summary_month_key(analysis)
        grouped_analyses.setdefault(month_key, []).append(analysis)

    results: list[tuple[Path, int, int, int, int]] = []
    for month_key in sorted(grouped_analyses):
        monthly_analyses = grouped_analyses[month_key]
        summary_path = order_summary_path(database_root, month_key)
        existing_headers, summary_rows = ([], {}) if reset else read_existing_order_summary(summary_path)
        base_headers = [header for header in existing_headers if header not in ORDER_SUMMARY_METADATA_HEADERS]
        incoming_headers: list[str] = []
        incoming_records: list[tuple[str, dict[str, str]]] = []

        for analysis in sorted(monthly_analyses, key=lambda item: (analysis_summary_datetime(item), item.target_path.name)):
            rows = read_rows(analysis.target_path, max_rows, max_cols)
            signature, header_row_index, headers, _matched = match_signature(rows, signatures)
            if signature is None or signature.report_type != "订单数据":
                continue
            order_index = find_header_index(headers, ORDER_ID_HEADERS)
            if order_index is None:
                continue
            append_headers(incoming_headers, headers)
            latest_download_time = download_datetime(analysis.target_path).strftime("%Y-%m-%d %H:%M:%S")
            for row in rows[header_row_index + 1 :]:
                if order_index >= len(row):
                    continue
                order_id = normalize_text(row[order_index])
                if not order_id:
                    continue
                record = row_to_dict(headers, row)
                record["汇总_店铺名称"] = analysis.shop_name
                record["汇总_最新下载时间"] = latest_download_time
                record["汇总_最新来源文件"] = analysis.source.name
                record["汇总_最新归档文件"] = analysis.target_path.name
                incoming_records.append((order_id, record))

        final_headers: list[str] = []
        append_headers(final_headers, base_headers)
        append_headers(final_headers, incoming_headers)
        append_headers(final_headers, ORDER_SUMMARY_METADATA_HEADERS)

        inserted = 0
        updated = 0
        for order_id, record in incoming_records:
            if order_id in summary_rows:
                updated += 1
                summary_rows[order_id] = merge_order_summary_record(summary_rows[order_id], record, final_headers)
            else:
                inserted += 1
                summary_rows[order_id] = {header: normalize_text(record.get(header, "")) for header in final_headers}

        output_rows = [
            {header: normalize_text(row.get(header, "")) for header in final_headers}
            for row in summary_rows.values()
        ]
        write_utf8_sig_csv(summary_path, output_rows, final_headers)
        results.append((summary_path, len(monthly_analyses), len(output_rows), inserted, updated))
    return results


def archived_order_analyses(database_root: Path, signatures: list[Signature], max_rows: int, max_cols: int) -> list[Analysis]:
    order_root = database_root / "订单数据"
    if not order_root.exists():
        return []

    analyses: list[Analysis] = []
    for path in sorted(order_root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS or is_order_summary_file(path):
            continue
        try:
            rows = read_rows(path, max_rows, max_cols)
            signature, header_row_index, headers, matched = match_signature(rows, signatures)
        except Exception:
            continue
        if signature is None or signature.report_type != "订单数据":
            continue

        file_download_date = download_date(path)
        period, period_source = decide_period(path, signature, rows, header_row_index, headers, file_download_date)
        shop_name = ""
        try:
            relative = path.relative_to(order_root)
            if relative.parts and not relative.parts[0].startswith("_"):
                shop_name = relative.parts[0]
        except ValueError:
            pass
        analyses.append(
            Analysis(
                source=path,
                status="ready",
                reason="archived_order",
                report_type="订单数据",
                target_folder="订单数据",
                shop_name=shop_name,
                shop_source="archive_path" if shop_name else "",
                period=period,
                period_source=period_source,
                download_date=file_download_date.isoformat(),
                target_path=path,
                product_counts="",
                matched_headers="|".join(matched),
                missing_fields="",
                map_updates=[],
                product_ids=[],
                new_product_ids=[],
            )
        )
    return analyses


def print_summary(rows: list[dict[str, str]], manifest_path: Path) -> None:
    counts = Counter(row["status"] for row in rows)
    print(f"manifest: {manifest_path}")
    print("summary:", ", ".join(f"{key}={value}" for key, value in sorted(counts.items())) or "none")
    for row in rows:
        print(f"[{row['status']}] {row['source_name']} -> {row['target_path']}")
        if row["reason"] != "ok":
            print(f"  reason: {row['reason']} missing={row['missing_fields']}")


def apply_actions(
    analyses: list[Analysis],
    database_root: Path,
    update_map: bool,
    product_map: dict[str, str],
    max_rows: int,
    max_cols: int,
) -> None:
    for analysis in analyses:
        if analysis.status != "ready":
            continue
        if analysis.report_type == ORDER_REPORT_TYPE:
            outputs = merge_order_data(analysis, database_root, max_rows, max_cols)
            for final_path, source_count, row_count in outputs:
                print(
                    "order data merged:",
                    f"{final_path}",
                    f"sources={source_count}",
                    f"rows={row_count}",
                )
            continue
        if analysis.report_type == AFTER_SALE_REPORT_TYPE:
            final_path, source_count, row_count = merge_after_sale_data(analysis, max_rows, max_cols)
            print(
                "after-sale data merged:",
                f"{final_path}",
                f"sources={source_count}",
                f"rows={row_count}",
            )
            continue
        if analysis.report_type == PROMOTION_ADJUSTMENT_REPORT_TYPE:
            final_path, source_count, row_count = merge_promotion_adjustment_log(analysis, max_rows, max_cols)
            print(
                "promotion adjustment log merged:",
                f"{final_path}",
                f"sources={source_count}",
                f"rows={row_count}",
            )
            continue
        target = analysis.target_path
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(analysis.source), str(target))
        if update_map:
            for product_id, shop_name in analysis.map_updates:
                product_map[product_id] = shop_name

    for analysis in analyses:
        if analysis.status != "duplicate_file":
            continue
        target = analysis.target_path
        if target.exists():
            target = unique_path(target)
            analysis.target_path = target
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(analysis.source), str(target))
        retained_name = analysis.reason.removeprefix("exact_duplicate_of:")
        print(f"exact duplicate moved: {analysis.source} -> {target}; retained={retained_name}")

    if update_map:
        save_product_map(product_map)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Organize shop backend downloaded reports.")
    parser.add_argument("--database-root", required=True, help="Operations database root folder.")
    parser.add_argument("--input", help="Input folder. Defaults to <database-root>\\下载未分类.")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Preview only. This is the default.")
    mode.add_argument("--apply", action="store_true", help="Move and rename files.")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=PREVIEW_MAX_ROWS,
        help=(
            "Maximum rows used for preview/classification (default: 100000). "
            "Formal order merging always streams and validates every row."
        ),
    )
    parser.add_argument("--max-cols", type=int, default=120, help="Maximum columns to inspect per file.")
    parser.add_argument("--no-update-map", action="store_true", help="Do not update shop_product_map.csv on apply.")
    parser.add_argument(
        "--allow-missing-required",
        action="store_true",
        help="Allow apply to continue when required report types are missing, for supplemental archive batches.",
    )
    parser.add_argument(
        "--rebuild-order-summary",
        action="store_true",
        help="Rebuild monthly 订单数据\\订单汇总表_YYYY年MM月.csv files from archived order reports without moving input files.",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    database_root = Path(args.database_root).resolve()
    input_dir = Path(args.input).resolve() if args.input else database_root / "下载未分类"
    apply = bool(args.apply)

    if not database_root.exists():
        print(f"database root does not exist: {database_root}", file=sys.stderr)
        return 2
    signatures = load_signatures()
    expected_reports = load_expected_reports()
    if args.rebuild_order_summary:
        analyses = archived_order_analyses(database_root, signatures, args.max_rows, args.max_cols)
        summary_results = update_order_summary(
            database_root, analyses, signatures, args.max_rows, args.max_cols, reset=True
        )
        if not summary_results:
            print("order summary rebuilt: no archived order reports found")
        for summary_path, report_count, total_orders, inserted_orders, updated_orders in summary_results:
            print(
                "order summary rebuilt:",
                f"{summary_path}",
                f"reports={report_count}",
                f"total={total_orders}",
                f"inserted={inserted_orders}",
                f"updated={updated_orders}",
            )
        return 0

    if not input_dir.exists():
        print(f"input folder does not exist: {input_dir}", file=sys.stderr)
        return 2

    try:
        extract_zip_inputs(input_dir)
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    product_map = load_product_map()
    archived_product_map = load_archived_product_map(database_root, signatures, product_map, args.max_rows, args.max_cols)
    product_map.update(archived_product_map)
    shops = known_shops(database_root, product_map)
    files = iter_input_files(input_dir)

    analyses = [
        analyze_file(path, database_root, signatures, product_map, shops, args.max_rows, args.max_cols)
        for path in files
    ]
    file_download_dates = {path: download_date(path) for path in files}
    apply_batch_shop_inference(database_root, analyses, signatures, file_download_dates)
    reserve_unique_targets(analyses)
    try:
        mark_exact_duplicate_inputs(analyses, input_dir)
    except OSError as exc:
        print(f"duplicate-file precheck failed: {exc}", file=sys.stderr)
        return 2

    preliminary_action = "apply" if apply else "dry-run"
    preliminary_reminder_rows = missing_report_reminder_rows(expected_reports, analyses, preliminary_action, input_dir)
    apply_blocked = (
        apply and has_missing_required_reports(preliminary_reminder_rows) and not args.allow_missing_required
    )
    action = "apply-blocked" if apply_blocked else preliminary_action
    rows = [manifest_row(analysis, action) for analysis in analyses]
    manifest_path = manifest_output_path(database_root, apply, apply_blocked)
    write_utf8_sig_csv(
        manifest_path,
        rows,
        [
            "action",
            "status",
            "reason",
            "source_path",
            "source_name",
            "target_path",
            "target_name",
            "report_type",
            "target_folder",
            "shop_name",
            "shop_source",
            "period",
            "period_source",
            "download_date",
            "product_counts",
            "matched_headers",
            "missing_fields",
            "map_updates",
            "new_product_ids",
        ],
    )
    reminder_path = missing_reminder_path(database_root, apply)
    reminder_rows = missing_report_reminder_rows(expected_reports, analyses, action, input_dir)
    write_utf8_sig_csv(reminder_path, reminder_rows, MISSING_REMINDER_FIELDNAMES)

    if apply_blocked:
        print_missing_reminders(reminder_rows, reminder_path)
        print("apply blocked: required reports are missing; no files were moved.")
        print_summary(rows, manifest_path)
        return 1

    if apply:
        try:
            validate_order_merge_inputs(analyses, database_root, args.max_cols)
            apply_actions(
                analyses,
                database_root=database_root,
                update_map=not args.no_update_map,
                product_map=product_map,
                max_rows=args.max_rows,
                max_cols=args.max_cols,
            )
        except RuntimeError as exc:
            print(f"apply stopped: {exc}", file=sys.stderr)
            return 2

    print_missing_reminders(reminder_rows, reminder_path)
    print_summary(rows, manifest_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
